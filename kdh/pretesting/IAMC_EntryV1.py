import sys
import atexit
from datetime import date, datetime
import re
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as xl
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QFileDialog, QCheckBox

from IAMC_Pretest_GUI import MyWindow, DictEditorApp

def true2yes(x):
    if x:
        return "Yes" 
    
    else: 
        return "No"

def exit_handler(workbook):
    try:
        workbook.close()
    except Exception as e:
        print("An exception occurred:", e)

def load_data(filepath):
    try:
        header_dataframe = pd.read_table(filepath, delimiter=',', nrows=0, header=0)
        df = pd.read_table(filepath, skiprows=1, usecols=[0, 1, 2, 3], header=0, index_col=False)
        return header_dataframe, df
    except FileNotFoundError:
        sys.exit(1)

def process_frequency_data(column_labels, df,band):
    #scales frequency and retrieves current with RF value
   
    current=[]
    
    freq_factor = int(column_labels[1].split()[0])
    scaled_freq = df.iloc[:, 0] * freq_factor
    power_mw = df.iloc[:, 1]
    
    mw2dbm = lambda x: 10 * np.log10(x)
    
    
    power_dbm=mw2dbm(power_mw)
    power_dbm=np.nan_to_num(power_dbm,nan=0.0)#why doesn't this work?
    power_in_band=[]
    freq_in_band=[]
    for i in range(len(scaled_freq)):
        if np.round(scaled_freq[i])>= band[0] and np.round(scaled_freq[i])<= band[1]:
            power_in_band.append(power_dbm[i])
            freq_in_band.append(scaled_freq[i])
    
   
    
    current.append(np.min(df.iloc[:,2]))
    current.append(np.max(df.iloc[:,2]))
    
    return scaled_freq, power_dbm,power_in_band,freq_in_band,current


def extract_info_main(headers):
    flat_headers = ''.join(headers)
    #date testing begins
    #date_tested, date_built = extract_dates(flat_headers)
    dates=extract_dates(flat_headers)
    date_tested=dates[len(dates)-1]
    date_built=dates[0]
    print(headers)
    print(flat_headers)
    #print (dates)
    #print (date_tested)
    #print (date_built)
    #date testing end
    tech = extract_tech(flat_headers)
    block = extract_block(flat_headers)
    pcb_rev, pcb_lot = extract_pcb_info(headers)
    tsc_lot=extract_tsc_lot_info(flat_headers,headers)
    return date_tested, date_built, tech, block, pcb_rev, pcb_lot,tsc_lot

def extract_dates(flat_headers):
    date_match = re.findall(r'[0-9]+/[0-9]+/[0-9]+', flat_headers, re.IGNORECASE)
    date_match.sort() #sort from earliest to oldest dates found
    if date_match:
        return date_match
    else:
        return ""
    #return date_match[0] if date_match else "", date_match[1] if len(date_match) > 1 else "" #syntax didn't work like i thought??

def extract_tech(flat_headers):
    #search for initials in header string
    tech_match = re.findall(r'\s[a-z]{3}\s', flat_headers, re.IGNORECASE)
    return tech_match[0] if tech_match and not (tech_match[0] == "SSP") else ""

def extract_block(flat_headers):
    block_match = re.search(r'R[0-9]{1}(\s+)B[0-9]{1}-[0-9]{2}([a-z]?)', flat_headers, re.IGNORECASE)
    return block_match.group(0) if block_match else ""

def extract_pcb_info(headers):
    pcb_rev, pcb_lot = "", ""
    indexes_pcb = [index for index, string in enumerate(headers) if ("pcb" in string.lower() or "pcbr*" in string.lower())]
    if indexes_pcb:
        pcb_rev, pcb_lot = process_pcb_info(headers[indexes_pcb[0]])
    return pcb_rev, pcb_lot

def process_pcb_info(pcb_info):
    pcb_rev_tmp = pcb_info.lstrip().rstrip()
    rev_match = re.search(r'pcb\s*r\s*(\d+)', pcb_rev_tmp.lower())
    pcb_rev = rev_match.group(1) if rev_match else ""
    pcblot_match = re.search(r'(\d{8})', pcb_rev_tmp.lower())
    pcb_lot = pcblot_match.group(1) if pcblot_match else ""
    return pcb_rev, pcb_lot

def extract_tsc_lot_info(flat_headers,headers):
    indexes_tsc = [index for index, string in enumerate(headers) if "tsc" in string.lower()]
    #indexes_tsc = [index for index, string in enumerate(headers) if "TSC" in string and( "lot" in string.lower())]
    if indexes_tsc:
        tsc_lot=headers[indexes_tsc[0]]
        tsc_lot=tsc_lot.lstrip()
        tsclot_match = re.search(r'[0-9]{8}',tsc_lot,re.IGNORECASE)
        if tsclot_match:
            tsc_lot=tsclot_match.group(0)
            return tsc_lot
    else:
        tsc_lot=""
        return tsc_lot

    lotsearch=re.split(r'\w\w-\w\w\w\w-\w\w\w\w',flat_headers,re.IGNORECASE)
    if (len(lotsearch)>1):
        lotsearch=lotsearch[1]
        lotmatch=re.search(r'(\s?)[0-9]{8}(\s?)',lotsearch,re.IGNORECASE)
        if lotmatch:
            tsc_lot=lotmatch.group(0)
            return tsc_lot
        
    
#waveguide bands
#pass to window application   
#need to update to inlcude rest of bands     
band_dict={}
band_dict["WR6.5"]=[110, 170]
band_dict["WR9.0"]=[82, 125]
band_dict["WR10.0"]=[75, 110]
band_dict["WR12.0"]=[60, 90]



###### main 
def main():
    # Run initial GUI
    app = QApplication(sys.argv)
    window = MyWindow(band_dict)
    window.show()
    app.exec_()

    # Grab user input information from GUI
    filepath = window.result_filepath
    test_notes = window.result_note
    tsc = window.result_TSC
    part_good = window.result_part
    band=window.result_value
    
    tracker_path = window.result_excel_filepath
    tracker_sheet = window.result_sheet

    # Load data from the file
    header_dataframe, df = load_data(filepath)

    # Grab column labels from data headers
    column_labels = df.columns.values.tolist()
    column_labels = column_labels[1]
    column_labels = column_labels.split(":")

    # Process frequency data
    
    
    scaled_freq, power_dbm,power_in_band,freq_in_band,current= process_frequency_data(column_labels, df,band)
    typicalPower=np.median(power_in_band)
    minPower=np.min(power_in_band)

   

    # Extract date, tech, block, and PCB info
    headers = header_dataframe.columns.values.tolist()
    date_tested, date_built, tech, block, pcb_rev, pcb_lot,tsc_lot = extract_info_main(headers)
    
 
    
    # Create header data export dictionary
    
    header_dict = {
        "Part Good": true2yes(part_good),
        "TSC Good": true2yes(tsc),
        "Type(M3/M6) etc": "",
        "Current (no RF)":"",
        "Max Current with RF":current[1],
        "Tech": tech,
        "Date Tested": date_tested,
        "Date Built": date_built,
        "Build": block[-1] if block and block[-1].isalpha() else 'A',
        "TSC Lot": tsc_lot,
        "Block": block,
        "PCB Rev": pcb_rev,
        "PCB Lot": pcb_lot,
        "Typical Power(dBm)":typicalPower,
        "Minimum Power(dBm)":minPower,
        "Data Path": filepath,
        "Notes": test_notes,
    }

    # Call of final data submittal screen
    
    #app = QApplication(sys.argv) # I BELIEVE THIS IS NOT NEEDEED. I REMOVED THIS LINE AND IT RUNS FINE???
    
    try:
        window = DictEditorApp(header_dict, scaled_freq, power_dbm,typicalPower,minPower)
    except:
        print("Tracking sheet is probably open")
        
        
    window.show()
    app.exec_()

    # Plot and save
    workbook = xl.load_workbook(tracker_path)
    workbook.active = workbook[tracker_sheet]
    worksheet = workbook.active
    row2use = len(worksheet['A'])

    # Must save after each write
    cell=worksheet.cell(row2use+1,1)
    cell.value=header_dict["Block"]
    try: 
        workbook.save(tracker_path)
    except:
        print("Tracking sheet is probably open")
    #workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,2)
    cell.value=header_dict["Build"]
    workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,3)
    cell.value=header_dict["Type(M3/M6) etc"]
    workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,8)
    cell.value=header_dict["Current (no RF)"]
    workbook.save(tracker_path)

    cell=worksheet.cell(row2use+1,9)
    cell.value=header_dict["Max Current with RF"]
    workbook.save(tracker_path)
 
 
    cell=worksheet.cell(row2use+1,4)
    cell.value=header_dict["Tech"]
    workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,5)
    cell.value=header_dict["Date Built"]
    workbook.save(tracker_path)
 
    #PCB rev currently not used in tracking doc. uncomment and change cell=worksheet.cell(row2use+1,7) to specify where to write in excel doc
    #cell=worksheet.cell(row2use+1,7)
    #cell.value=header_dict["PCB Rev"]
    #workbook.save(tracker_path)
 
 
    cell=worksheet.cell(row2use+1,6)# last used row and column 
    cell.value=header_dict["TSC Lot"]
    workbook.save(tracker_path)
    
    cell=worksheet.cell(row2use+1,13)# last used row and column 
    cell.value=header_dict["Typical Power(dBm)"]
    workbook.save(tracker_path)
    
    #Minimum power currently not used in tracking doc
    #cell=worksheet.cell(row2use+1,10)# last used row and column 
    #cell.value=header_dict["Minimum Power(dBm)"]
    #workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,15)# last used row and column 
    cell.value=header_dict["Part Good"]
    workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,16)# last used row and column 
    cell.value=header_dict["TSC Good"]
    workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,17)# last used row and column 
    cell.value=header_dict["Notes"]
    workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,18)# last used row and column 
    cell.value=header_dict["Date Tested"]
    workbook.save(tracker_path)
 
    cell=worksheet.cell(row2use+1,19)# last used row and column 
    cell.hyperlink=header_dict["Data Path"]
    workbook.save(tracker_path)


    # (Add other cell updates here)

    
    
    ## Close writer or Excel file stays locked for editing
    workbook.close()
    main()
    #sys.exit()

if __name__ == "__main__":
    main()
